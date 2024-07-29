

'''
Decorators
'''

class Decotators():
    def timer(func):
        def func_wrapper(*args, **kwargs):
            from time import time   ##import The Time Module
            time_start = time()
            result = func(*args, **kwargs)
            time_end = time()
            time_spend = time_end - time_start
            print('%s cost time: %.4f s' % (func.__name__, time_spend))
            return result
        return func_wrapper
'''
ToolKit
'''
class Tools():
    """change csv into excel
    Args:
            sourcePath:str The path of CSV file
            savePath:str The path you want to Save Excels
            encode='utf-8' Encoding form select
    """
    def csv_to_xlsx_xl(self,sourcePath:str,savePath:str,encode='utf-8'):
        from openpyxl import Workbook   ##import the External packages
        
        f = open(sourcePath, 'r', encoding=encode)
        # Create a workbook
        workbook = Workbook()
        # Create a worksheet
        worksheet = workbook.active
        workbook.title = 'sheet'

        for line in f:
            row = line.split(',')
            worksheet.append(row)
        workbook.save(savePath)
    
    '''
    Args:
    path --> the Expected Dir you want to make
    Flag_Output --> The Log Output Situation    Default : False
            Ture --> put something into terminal
            False --> put nothing
    '''
    def Make_dir_Safe(Path:str,Name:str,Flag_Output:bool):
        from os import path,mkdir
        if bool(path.exists(Path)) == False:##this maens IF the dir not exists
            try:
                mkdir(Name)
            except:
                if Flag_Output:
                    print("{0} had Existed!".format(Path))
            
    
    '''
    Args:
        Excel_path --> Excel File Path
        Json_path --> Json File Path
        encode --> The encode of the json File
    '''
    def excel_to_json_xl(Excel_path:str, Json_path:str,encode:str='utf-8'):
        # 导入必要的库
        import openpyxl
        import json
        # 加载Excel文件
        wb = openpyxl.load_workbook(Excel_path)
        # 选择第一个工作表
        sheet = wb.active
        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        # 创建一个空列表来存储数据
        data = []
        # 遍历工作表中的每一行，跳过表头
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 将每一行数据转换为字典，键为表头，值为单元格值
            row_data = dict(zip(headers, row))
            # 将字典添加到数据列表中
            data.append(row_data)
        # 将数据转换为JSON格式
        json_data = json.dumps(data, ensure_ascii=False, indent=4)
        # 将JSON数据写入文件
        with open(Json_path, 'w', encoding=encode) as f:
            f.write(json_data)

        

    def get_filenames(directory:str):
        import os
        # 获取目录下的所有文件名
        filenames = os.listdir(directory)
        # 过滤出文件名，不包括子目录
        filenames = [filename for filename in filenames if os.path.isfile(os.path.join(directory, filename))]
        # 返回文件名列表
        return filenames
    


'''
If your Editer Supports Autocomplete Way, Don't use this 
This is for the Unmodified Vim or other editer
'''
class Abbr_():
    def MakeDir_s(Path:str):
        Tools.Make_dir_Safe(Path)

    def csv2xlsx(sourcePath:str, savePath:str):
        Tools.csv_to_xlsx_xl(sourcePath, savePath)
    
    def xlsx2json(sourcePath:str, savePath:str):
        Tools.excel_to_json_xl(sourcePath, savePath)
    
    def get_names(directory:str):
        return Tools.get_filenames(directory)

    